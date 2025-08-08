import io
import os
import glob
from datetime import datetime
from typing import Dict, List, Tuple
import pandas as pd
import streamlit as st
import altair as alt
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="Excel → App Web (Persistente)", layout="wide")

# Pasta para salvar arquivos
DATA_DIR = os.getenv("DATA_DIR", "./data")
os.makedirs(DATA_DIR, exist_ok=True)

def save_uploaded_to_data_dir(uploaded_file) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = uploaded_file.name.replace("/", "_").replace("\\", "_")
    out_path = os.path.join(DATA_DIR, f"{ts}__{safe_name}")
    with open(out_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    return out_path

@st.cache_data(show_spinner=False)
def read_all_sheets(file_bytes: bytes) -> Tuple[Dict[str, pd.DataFrame], bytes]:
    return pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, engine="openpyxl"), file_bytes

def load_workbook_keep_vba(file_bytes: bytes):
    return load_workbook(io.BytesIO(file_bytes), keep_vba=True, data_only=False)

def write_df_to_ws(ws, df: pd.DataFrame):
    ws.delete_rows(1, ws.max_row)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

# ===================== Sidebar =====================
st.sidebar.title("📄 Arquivo (Persistente)")
uploaded_file = st.sidebar.file_uploader("Envie um Excel (.xlsm ou .xlsx)", type=["xlsm", "xlsx"])

# Lista de arquivos salvos
saved_files = sorted(glob.glob(os.path.join(DATA_DIR, "*.xlsm")), key=os.path.getmtime, reverse=True)
chosen = st.sidebar.selectbox("Arquivos salvos", saved_files) if saved_files else None

file_bytes = None
if uploaded_file:
    path = save_uploaded_to_data_dir(uploaded_file)
    with open(path, "rb") as f:
        file_bytes = f.read()
elif chosen:
    with open(chosen, "rb") as f:
        file_bytes = f.read()

if not file_bytes:
    st.info("Envie ou selecione um arquivo.")
    st.stop()

# ===================== Leitura do Excel =====================
sheets, raw = read_all_sheets(file_bytes)
st.session_state.setdefault("sheets", {k: v.copy() for k, v in sheets.items()})
st.session_state["original_bytes"] = raw

# ===================== Interface =====================
tabs = st.tabs(["📝 Edição", "📊 Relatórios", "📈 Gráficos", "💾 Salvar"])

# Aba de Edição
with tabs[0]:
    sel = st.selectbox("Aba", list(st.session_state["sheets"].keys()))
    edited = st.data_editor(st.session_state["sheets"][sel], num_rows="dynamic", use_container_width=True)
    if st.button("Salvar edição desta aba"):
        st.session_state["sheets"][sel] = edited.copy()
        st.success("Alterações aplicadas.")

# Aba de Relatórios
with tabs[1]:
    aba_rel = st.selectbox("Aba base", list(st.session_state["sheets"].keys()))
    df_rel = st.session_state["sheets"][aba_rel]
    cols = list(df_rel.columns)
    group_cols = st.multiselect("Agrupar por", options=cols)
    num_cols = [c for c in cols if pd.api.types.is_numeric_dtype(df_rel[c])]
    value_col = st.selectbox("Coluna de valores", options=(num_cols or cols))
    agg = st.selectbox("Agregação", ["sum", "mean", "count", "min", "max"])
    if st.button("Gerar relatório"):
        if agg == "count":
            rpt = df_rel.groupby(group_cols).size().reset_index(name="count")
        else:
            rpt = df_rel.groupby(group_cols).agg({value_col: agg}).reset_index()
        st.dataframe(rpt, use_container_width=True)

# Aba de Gráficos
with tabs[2]:
    aba_graf = st.selectbox("Aba base", list(st.session_state["sheets"].keys()))
    df_graf = st.session_state["sheets"][aba_graf]
    if len(df_graf.columns) >= 2:
        x = st.selectbox("Eixo X", options=df_graf.columns)
        y = st.selectbox("Eixo Y", options=[c for c in df_graf.columns if c != x])
        chart_type = st.radio("Tipo", ["Linha", "Barra", "Área", "Dispersão"], horizontal=True)
        try:
            df_graf[y] = pd.to_numeric(df_graf[y])
        except:
            pass
        if chart_type == "Linha":
            ch = alt.Chart(df_graf).mark_line().encode(x=x, y=y)
        elif chart_type == "Barra":
            ch = alt.Chart(df_graf).mark_bar().encode(x=x, y=y)
        elif chart_type == "Área":
            ch = alt.Chart(df_graf).mark_area().encode(x=x, y=y)
        else:
            ch = alt.Chart(df_graf).mark_circle().encode(x=x, y=y)
        st.altair_chart(ch, use_container_width=True)

# Aba de Salvar
with tabs[3]:
    out_name = st.text_input("Nome do arquivo", value="saida_atualizada.xlsm")
    if st.button("💾 Salvar arquivo Excel (.xlsm)"):
        wb = load_workbook_keep_vba(st.session_state["original_bytes"])
        for name, df in st.session_state["sheets"].items():
            if name not in wb.sheetnames:
                ws = wb.create_sheet(name)
            else:
                ws = wb[name]
            write_df_to_ws(ws, df)
        bio_out = io.BytesIO()
        wb.save(bio_out)
        st.download_button("⬇️ Baixar .xlsm", bio_out.getvalue(), file_name=out_name)
        # Salvar no diretório persistente
        dest = os.path.join(DATA_DIR, out_name)
        with open(dest, "wb") as f:
            f.write(bio_out.getvalue())
