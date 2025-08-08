
import io
import os
import time
from typing import Dict, List, Tuple

import pandas as pd
import streamlit as st
import altair as alt
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="Excel → Web App", layout="wide")

# -------------------------
# Helpers
# -------------------------
@st.cache_data(show_spinner=False)
def read_all_sheets(file_bytes: bytes) -> Tuple[Dict[str, pd.DataFrame], bytes]:
    """
    Read all sheets from an .xlsm/.xlsx file into a dict of DataFrames.
    Also return the original bytes (to preserve macros when saving with keep_vba).
    """
    return pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, engine="openpyxl"), file_bytes

def load_workbook_keep_vba(file_bytes: bytes):
    """Load workbook with keep_vba to preserve macros when saving."""
    bio = io.BytesIO(file_bytes)
    wb = load_workbook(bio, keep_vba=True, data_only=False)
    return wb

def df_safe_cast(df: pd.DataFrame) -> pd.DataFrame:
    """Try to reduce dtype issues after editing in Streamlit."""
    for col in df.columns:
        # If column is object but looks like number, try cast
        if df[col].dtype == "object":
            try:
                df[col] = pd.to_numeric(df[col])
            except Exception:
                pass
        # Datetime handling
        if "date" in str(col).lower():
            try:
                df[col] = pd.to_datetime(df[col])
            except Exception:
                pass
    return df

def write_df_to_ws(ws, df: pd.DataFrame):
    """Overwrite a worksheet with DataFrame content (values only)."""
    # Clear existing data (values only) by deleting rows except headers
    ws.delete_rows(1, ws.max_row)
    # Write header + rows
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

def make_pivot(df: pd.DataFrame, by_cols: List[str], value_col: str, aggfunc: str) -> pd.DataFrame:
    if not by_cols or not value_col:
        return pd.DataFrame()
    agg_map = {value_col: aggfunc}
    pt = df.groupby(by_cols, dropna=False).agg(agg_map).reset_index()
    pt = pt.rename(columns={value_col: f"{aggfunc}({value_col})"})
    return pt

def chart_from_df(df: pd.DataFrame, x: str, y: str, color: str, chart_type: str):
    base = alt.Chart(df).mark_point().encode()
    if chart_type == "Linha":
        chart = alt.Chart(df).mark_line().encode(x=x, y=y, color=color if color else alt.value(None), tooltip=list(df.columns))
    elif chart_type == "Barra":
        chart = alt.Chart(df).mark_bar().encode(x=x, y=y, color=color if color else alt.value(None), tooltip=list(df.columns))
    elif chart_type == "Área":
        chart = alt.Chart(df).mark_area().encode(x=x, y=y, color=color if color else alt.value(None), tooltip=list(df.columns))
    elif chart_type == "Dispersão":
        chart = alt.Chart(df).mark_circle().encode(x=x, y=y, color=color if color else alt.value(None), tooltip=list(df.columns))
    else:
        chart = alt.Chart(df).mark_line().encode(x=x, y=y, color=color if color else alt.value(None), tooltip=list(df.columns))
    return chart.properties(use_container_width=True)

def init_session_state(sheet_dfs: Dict[str, pd.DataFrame]):
    if "sheets" not in st.session_state:
        st.session_state["sheets"] = {name: df.copy() for name, df in sheet_dfs.items()}
    if "original_bytes" not in st.session_state:
        st.session_state["original_bytes"] = None

# -------------------------
# Sidebar - file handling
# -------------------------
st.sidebar.title("📄 Arquivo")
uploaded_file = st.sidebar.file_uploader("Envie o Excel (.xlsm ou .xlsx)", type=["xlsm", "xlsx"], accept_multiple_files=False)

# Load default file if exists in working dir
DEFAULT_FILENAME = "Strategic_Plan_2025_Rev01.xlsm"
default_path = os.path.join(".", DEFAULT_FILENAME)

if uploaded_file is not None:
    file_bytes = uploaded_file.read()
    try:
        all_sheets, raw = read_all_sheets(file_bytes)
        init_session_state(all_sheets)
        st.session_state["original_bytes"] = raw
        st.sidebar.success(f"Carregado: {uploaded_file.name} com {len(all_sheets)} abas.")
    except Exception as e:
        st.sidebar.error(f"Falha ao ler: {e}")
elif os.path.exists(default_path):
    with open(default_path, "rb") as f:
        file_bytes = f.read()
    all_sheets, raw = read_all_sheets(file_bytes)
    init_session_state(all_sheets)
    st.session_state["original_bytes"] = raw
    st.sidebar.info(f"Usando arquivo padrão: {DEFAULT_FILENAME} ({len(all_sheets)} abas).")
else:
    st.info("Envie um arquivo Excel (.xlsm/.xlsx) na barra lateral para começar.")
    st.stop()

sheets: Dict[str, pd.DataFrame] = st.session_state["sheets"]

# -------------------------
# Main UI
# -------------------------
st.title("🧩 Excel → App Web (Edição, Relatórios e Gráficos)")
st.caption("Edite dados, gere relatórios e crie gráficos interativos a partir de todas as abas do Excel.")

tabs = st.tabs(["📝 Edição", "📊 Relatórios", "📈 Gráficos", "💾 Salvar / Exportar"])

# -------------------------
# Tab 1: Edição
# -------------------------
with tabs[0]:
    st.subheader("Edição de Dados")
    sheet_names = list(sheets.keys())
    sel = st.selectbox("Escolha a aba para editar", sheet_names, key="edit_sheet_sel")
    df = sheets[sel]

    st.markdown("Use a tabela abaixo para **editar** valores, **adicionar** ou **excluir** linhas.")

    edited = st.data_editor(
        df,
        num_rows="dynamic",
        use_container_width=True,
        key=f"editor_{sel}",
        hide_index=True,
    )
    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("Aplicar alterações nesta aba", type="primary"):
            sheets[sel] = df_safe_cast(edited.copy())
            st.success(f"Alterações aplicadas na aba **{sel}**.")
    with col2:
        if st.button("Desfazer alterações não salvas (recarregar aba)"):
            # Recarrega do arquivo original
            try:
                fresh_sheets, _ = read_all_sheets(st.session_state["original_bytes"])
                sheets[sel] = fresh_sheets[sel].copy()
                st.experimental_rerun()
            except Exception as e:
                st.error(f"Erro ao recarregar: {e}")
    with col3:
        st.download_button(
            label="Baixar esta aba (CSV)",
            data=sheets[sel].to_csv(index=False).encode("utf-8"),
            file_name=f"{sel}.csv",
            mime="text/csv",
        )

# -------------------------
# Tab 2: Relatórios
# -------------------------
with tabs[1]:
    st.subheader("Construtor de Relatórios (tipo Tabela Dinâmica)")
    r_sel = st.selectbox("Escolha a aba base", list(sheets.keys()), key="report_sheet_sel")
    rdf = sheets[r_sel]

    cols = list(rdf.columns)
    group_cols = st.multiselect("Agrupar por (linhas)", options=cols)
    numeric_cols = [c for c in cols if pd.api.types.is_numeric_dtype(rdf[c])]
    value_col = st.selectbox("Medida (coluna numérica)", options=(numeric_cols or cols))
    aggfunc = st.selectbox("Agregação", options=["sum", "mean", "median", "min", "max", "count"])

    if st.button("Gerar relatório", type="primary"):
        if aggfunc == "count":
            rpt = rdf.groupby(group_cols, dropna=False).size().reset_index(name="count")
        else:
            rpt = make_pivot(rdf, group_cols, value_col, aggfunc)
        if rpt.empty:
            st.warning("Selecione pelo menos uma coluna para agrupar e uma medida.")
        else:
            st.dataframe(rpt, use_container_width=True)
            st.download_button(
                "Baixar relatório (CSV)",
                rpt.to_csv(index=False).encode("utf-8"),
                file_name=f"relatorio_{r_sel}.csv",
                mime="text/csv",
            )

# -------------------------
# Tab 3: Gráficos
# -------------------------
with tabs[2]:
    st.subheader("Gráficos Interativos")
    g_sel = st.selectbox("Escolha a aba base", list(sheets.keys()), key="chart_sheet_sel")
    gdf = sheets[g_sel]

    if gdf.empty or len(gdf.columns) < 2:
        st.warning("A aba selecionada não possui colunas suficientes para gráficos.")
    else:
        x = st.selectbox("Eixo X", options=list(gdf.columns), index=0)
        y = st.selectbox("Eixo Y", options=[c for c in gdf.columns if c != x], index=min(1, len(gdf.columns)-1))
        color = st.selectbox("Cor (opcional)", options=[""] + [c for c in gdf.columns if c not in [x, y]], index=0)
        chart_type = st.radio("Tipo de gráfico", options=["Linha", "Barra", "Área", "Dispersão"], horizontal=True)

        # Tentativa de conversão numérica para Y
        try:
            gdf[y] = pd.to_numeric(gdf[y])
        except Exception:
            pass

        ch = chart_from_df(gdf, x, y, color if color else None, chart_type)
        st.altair_chart(ch, use_container_width=True)

# -------------------------
# Tab 4: Save / Export
# -------------------------
with tabs[3]:
    st.subheader("Salvar de volta no Excel (preserva macros)")
    st.caption("Observação: formatações/estilos podem ser perdidos; macros (.xlsm) são preservadas.")

    colA, colB = st.columns(2)
    with colA:
        out_name = st.text_input("Nome do arquivo de saída", value="saida_atualizada.xlsm")
    with colB:
        include_only_selected = st.toggle("Incluir apenas a aba selecionada na Edição", value=False)

    if st.button("💾 Salvar arquivo Excel (.xlsm)"):
        try:
            wb = load_workbook_keep_vba(st.session_state["original_bytes"])
            # Decide quais abas salvar
            save_sheets = [st.session_state.get("edit_sheet_sel")] if include_only_selected else list(sheets.keys())

            for sheet_name in save_sheets:
                if sheet_name not in wb.sheetnames:
                    # criar se não existir
                    ws = wb.create_sheet(sheet_name)
                else:
                    ws = wb[sheet_name]
                write_df_to_ws(ws, sheets[sheet_name])

            bio_out = io.BytesIO()
            wb.save(bio_out)
            st.success("Arquivo salvo em memória. Use o botão abaixo para baixar.")

            st.download_button(
                label="⬇️ Baixar arquivo .xlsm",
                data=bio_out.getvalue(),
                file_name=out_name,
                mime="application/vnd.ms-excel.sheet.macroEnabled.12",
            )
        except Exception as e:
            st.error(f"Erro ao salvar: {e}")

    st.divider()
    st.subheader("Exportar todas as abas como CSVs")
    if st.button("Gerar ZIP com CSVs"):
        import zipfile
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
            for name, df in sheets.items():
                zipf.writestr(f"{name}.csv", df.to_csv(index=False))
        st.download_button(
            "⬇️ Baixar ZIP de CSVs",
            data=zip_buffer.getvalue(),
            file_name="todas_as_abas.zip",
            mime="application/zip",
        )

st.sidebar.markdown("---")
st.sidebar.caption("Dica: após editar, use a aba **Salvar/Exportar** para baixar o .xlsm preservando macros.")
